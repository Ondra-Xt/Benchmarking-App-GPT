# src/excel_export.py
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional
import json
import re

import pandas as pd
import openpyxl

_ILLEGAL_EXCEL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_ILLEGAL_ESCAPED_UNICODE_RE = re.compile(r"\\u00(?:0[0-8BCEFbcef]|1[0-9A-Fa-f])")
_EXCEL_MAX_CELL_LEN = 32767

BENCHMARK_SCORING_KEYS = [
    "flow_rate_score",
    "material_v4a_score",
    "din_en_1253_score",
    "din_en_18534_score",
    "height_adjustability_score",
    "sales_price_score",
    "outlet_flexibility_score",
    "sealing_fleece_score",
    "colour_count_score",
]

DEFAULT_BENCHMARK_WEIGHTS_PCT = {
    "flow_rate_score": 25,
    "material_v4a_score": 15,
    "din_en_1253_score": 10,
    "din_en_18534_score": 10,
    "height_adjustability_score": 10,
    "sales_price_score": 15,
    "outlet_flexibility_score": 5,
    "sealing_fleece_score": 5,
    "colour_count_score": 5,
}

LEGACY_EQUIVALENCE_KEYS = [
    "length_mode_match",
    "selected_length_match",
    "length_delta_within_tolerance",
    "equiv_finish_set_requires_base",
    "equiv_complete_system_bonus",
]


def _is_nan(x: Any) -> bool:
    try:
        return x != x  # NaN != NaN
    except Exception:
        return False


def _sanitize_excel_string(value: str) -> str:
    s = value or ""
    s = _ILLEGAL_EXCEL_XML_CHARS_RE.sub("", s)
    s = _ILLEGAL_ESCAPED_UNICODE_RE.sub("", s)

    if len(s) > _EXCEL_MAX_CELL_LEN:
        s = s[:_EXCEL_MAX_CELL_LEN]

    return s


def _to_excel_cell(v: Any) -> Any:
    """
    Převede hodnotu z DataFrame na hodnotu, kterou openpyxl umí uložit do buňky.

    - None/NaN -> None
    - list/tuple/set/dict -> JSON string
    - Path -> str
    - numpy scalar -> item()
    """
    if v is None or _is_nan(v):
        return None

    if hasattr(v, "item") and callable(getattr(v, "item")):
        try:
            v = v.item()
        except Exception:
            pass

    if isinstance(v, Path):
        return _sanitize_excel_string(str(v))

    if isinstance(v, (bytes, bytearray)):
        try:
            return _sanitize_excel_string(v.decode("utf-8", errors="ignore"))
        except Exception:
            return _sanitize_excel_string(str(v))

    if isinstance(v, (list, tuple, set, dict)):
        try:
            if isinstance(v, set):
                v = sorted(list(v))
            return _sanitize_excel_string(json.dumps(v, ensure_ascii=False))
        except Exception:
            return _sanitize_excel_string(str(v))

    if isinstance(v, str):
        return _sanitize_excel_string(v)

    return v




def _extract_source_checks(evidence_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = [
        "manufacturer", "source_id", "family", "source_url", "source_type", "status_code",
        "final_url", "content_hash_sha256", "content_length", "baseline_hash_sha256",
        "baseline_content_length", "hash_changed", "length_changed", "expected_terms_found",
        "expected_terms_missing", "new_source_candidate_count", "sample_new_source_candidates",
        "review_required", "review_reason", "checked_at", "extraction_mode", "fetch_error",
    ]
    if evidence_df.empty or "label" not in evidence_df.columns or "snippet" not in evidence_df.columns:
        return pd.DataFrame(columns=required_cols)
    rows = []
    for _, ev in evidence_df.iterrows():
        label = str(ev.get("label") or "")
        if not label.startswith("source_check:"):
            continue
        snippet = ev.get("snippet")
        try:
            payload = json.loads(snippet) if isinstance(snippet, str) else {}
        except Exception:
            payload = {}
        row = {k: payload.get(k, "") for k in required_cols}
        if not row.get("manufacturer"):
            row["manufacturer"] = str(ev.get("manufacturer") or "")
        if not row.get("source_id"):
            row["source_id"] = label.split(":", 1)[-1]
        rows.append(row)
    return pd.DataFrame(rows, columns=required_cols)


def _present(v: Any) -> bool:
    if v is None or _is_nan(v):
        return False
    s = str(v).strip().lower()
    return s not in {"", "nan", "none", "null", "unknown", "not_applicable"}




def _cfg_get(cfg: Any, key: str, default: Any = None) -> Any:
    if cfg is None:
        return default
    if hasattr(cfg, "get") and callable(getattr(cfg, "get")):
        try:
            return cfg.get(key, default)
        except Exception:
            pass
    return getattr(cfg, key, default)


def _extract_final_scoring_weights(cfg: Any) -> pd.DataFrame:
    raw = _cfg_get(cfg, "final_weights_pct", {}) or {}
    if not isinstance(raw, dict):
        raw = {}

    rows = []
    for key in BENCHMARK_SCORING_KEYS:
        val = raw.get(key, DEFAULT_BENCHMARK_WEIGHTS_PCT[key])
        try:
            val = float(val)
        except Exception:
            val = float(DEFAULT_BENCHMARK_WEIGHTS_PCT[key])
        rows.append({
            "key": key,
            "weight_pct": val,
            "enabled": True,
            "scoring_model": "benchmark_scoring_v2",
            "note": "active benchmark scoring criterion",
        })

    return pd.DataFrame(rows, columns=["key", "weight_pct", "enabled", "scoring_model", "note"])


def _extract_legacy_equivalence_weights(cfg: Any) -> pd.DataFrame:
    raw = _cfg_get(cfg, "equivalence_weights_pct", {}) or {}
    if not isinstance(raw, dict):
        raw = {}

    rows = []
    for key in LEGACY_EQUIVALENCE_KEYS:
        try:
            legacy_val = float(raw.get(key, 0) or 0)
        except Exception:
            legacy_val = 0.0
        rows.append({
            "key": key,
            "weight_pct": 0.0,
            "legacy_weight_pct": legacy_val,
            "enabled": False,
            "scoring_model": "legacy_equivalence_diagnostic_only",
            "note": "disabled; not used by benchmark_scoring_v2 final_score",
        })

    return pd.DataFrame(rows, columns=["key", "weight_pct", "legacy_weight_pct", "enabled", "scoring_model", "note"])


def _extract_config_sheet() -> pd.DataFrame:
    rows = [
        {"key": "active_scoring_model", "value": "benchmark_scoring_v2", "note": "active scoring model"},
        {"key": "benchmark_scoring_weights_sheet", "value": "Final_Scoring_Weights", "note": "active benchmark scoring weights"},
        {"key": "legacy_equivalence_weights_sheet", "value": "Legacy_Equivalence_Weights", "note": "disabled diagnostic-only legacy weights"},
    ]
    return pd.DataFrame(rows, columns=["key", "value", "note"])
def _scoring_field_coverage(products_df: pd.DataFrame, comparison_df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "manufacturer",
        "product_id",
        "product_name",
        "candidate_type",
        "complete_system",
        "in_products",
        "in_comparison",
        "has_flow_rate_lps",
        "has_material_data",
        "has_din_en_1253_data",
        "has_din_en_18534_data",
        "has_height_adjustability_data",
        "has_price_data",
        "has_outlet_flexibility_data",
        "has_sealing_fleece_data",
        "has_colour_count_data",
        "present_scoring_fields",
        "missing_scoring_fields",
        "scoring_readiness_pct",
        "scoring_readiness_note",
    ]

    products_df = pd.DataFrame() if products_df is None else products_df.copy()
    comparison_df = pd.DataFrame() if comparison_df is None else comparison_df.copy()

    def _present(value: Any) -> bool:
        if value is None:
            return False
        if _is_nan(value):
            return False
        text = str(value).strip().lower()
        if text in {"", "nan", "none", "null", "unknown", "not_applicable", "n/a", "na"}:
            return False
        return True

    def _text(r: pd.Series, keys: tuple[str, ...]) -> str:
        return " ".join(str(r.get(k) or "") for k in keys).lower()

    def _truthy(value: Any) -> bool:
        return str(value or "").strip().lower() in {"yes", "true", "1", "y", "ja"}

    def _has_positive_number(r: pd.Series, key: str) -> bool:
        try:
            value = pd.to_numeric(r.get(key), errors="coerce")
            return pd.notna(value) and float(value) > 0
        except Exception:
            return False

    def _has_height_adjustability(r: pd.Series) -> bool:
        try:
            hmin = pd.to_numeric(r.get("height_adj_min_mm"), errors="coerce")
            hmax = pd.to_numeric(r.get("height_adj_max_mm"), errors="coerce")
            return pd.notna(hmin) and pd.notna(hmax) and float(hmax) > float(hmin)
        except Exception:
            return False

    frames = []
    if not products_df.empty:
        p = products_df.copy()
        p["in_products"] = True
        p["in_comparison"] = False
        frames.append(p)

    if not comparison_df.empty:
        c = comparison_df.copy()
        c["in_products"] = False
        c["in_comparison"] = True
        frames.append(c)

    if not frames:
        return pd.DataFrame(columns=cols)

    all_rows = pd.concat(frames, ignore_index=True, sort=False)

    if "manufacturer" not in all_rows.columns:
        all_rows["manufacturer"] = ""
    if "product_id" not in all_rows.columns:
        all_rows["product_id"] = ""

    all_rows["manufacturer"] = all_rows["manufacturer"].astype(str)
    all_rows["product_id"] = all_rows["product_id"].astype(str)

    all_rows = all_rows.sort_values(
        by=["in_comparison", "in_products"],
        ascending=[False, False],
    )

    all_rows = all_rows.drop_duplicates(
        subset=["manufacturer", "product_id"],
        keep="first",
    )

    # Membership flags must reflect real membership in Products / Comparison,
    # not only the row that survived drop_duplicates.
    product_keys = set()
    if not products_df.empty and {"manufacturer", "product_id"}.issubset(products_df.columns):
        product_keys = set(
            zip(
                products_df["manufacturer"].astype(str),
                products_df["product_id"].astype(str),
            )
        )

    comparison_keys = set()
    if not comparison_df.empty and {"manufacturer", "product_id"}.issubset(comparison_df.columns):
        comparison_keys = set(
            zip(
                comparison_df["manufacturer"].astype(str),
                comparison_df["product_id"].astype(str),
            )
        )

    out = []

    groups = [
        (
            "has_flow_rate_lps",
            lambda r: _has_positive_number(r, "flow_rate_lps"),
        ),
        (
            "has_material_data",
            lambda r: any(
                _present(r.get(k))
                for k in ("material_v4a", "material_detail", "material_class")
            ),
        ),
        (
            "has_din_en_1253_data",
            lambda r: (
                "din en 1253"
                in _text(
                    r,
                    (
                        "din_en_1253",
                        "certification_din_en_1253",
                        "din_en_1253_cert",
                        "certifications",
                        "certificate_text",
                    ),
                )
                or _truthy(r.get("din_en_1253"))
                or _truthy(r.get("din_en_1253_cert"))
            ),
        ),
        (
            "has_din_en_18534_data",
            lambda r: (
                "din en 18534"
                in _text(
                    r,
                    (
                        "din_en_18534",
                        "certification_din_en_18534",
                        "din_18534_compliance",
                        "waterproofing_standard",
                        "certifications",
                        "certificate_text",
                    ),
                )
                or _truthy(r.get("din_en_18534"))
                or _truthy(r.get("din_18534_compliance"))
            ),
        ),
        (
            "has_height_adjustability_data",
            lambda r: _has_height_adjustability(r),
        ),
        (
            "has_price_data",
            lambda r: any(
                _has_positive_number(r, k)
                for k in ("sales_price", "sales_price_eur", "price_eur", "offer_price")
            ),
        ),
        (
            "has_outlet_flexibility_data",
            lambda r: any(
                _present(r.get(k))
                for k in (
                    "vertical_outlet_available",
                    "side_outlet_available",
                    "outlet_orientation",
                    "outlet_options",
                )
            ),
        ),
        (
            "has_sealing_fleece_data",
            lambda r: any(
                _present(r.get(k))
                for k in (
                    "sealing_fleece_preassembled",
                    "sealing_fleece",
                    "waterproofing_fleece_preassembled",
                )
            ),
        ),
        (
            "has_colour_count_data",
            lambda r: any(
                _present(r.get(k))
                for k in (
                    "colours_count",
                    "color_count",
                    "available_colours",
                    "available_colors",
                    "finish_count",
                )
            ),
        ),
    ]

    for _, r in all_rows.iterrows():
        key = (str(r.get("manufacturer") or ""), str(r.get("product_id") or ""))
        flags = {name: bool(fn(r)) for name, fn in groups}

        present = [name for name, value in flags.items() if value]
        missing = [name for name, value in flags.items() if not value]

        rec = {
            c: r.get(c, "")
            for c in (
                "manufacturer",
                "product_id",
                "product_name",
                "candidate_type",
                "complete_system",
            )
        }

        rec["in_products"] = key in product_keys
        rec["in_comparison"] = key in comparison_keys
        rec.update(flags)
        rec["present_scoring_fields"] = ",".join(present)
        rec["missing_scoring_fields"] = ",".join(missing)
        rec["scoring_readiness_pct"] = round((len(present) / len(groups)) * 100.0, 1)
        rec["scoring_readiness_note"] = (
            "ready" if len(missing) == 0 else f"missing:{len(missing)}"
        )

        out.append(rec)

    return pd.DataFrame(out, columns=cols)


def export_excel(
    template_path: str,
    out_path: str,
    cfg: Any,
    registry_df: Optional[pd.DataFrame] = None,
    products_df: Optional[pd.DataFrame] = None,
    comparison_df: Optional[pd.DataFrame] = None,
    excluded_df: Optional[pd.DataFrame] = None,
    evidence_df: Optional[pd.DataFrame] = None,
    bom_options_df: Optional[pd.DataFrame] = None,
    components_df: Optional[pd.DataFrame] = None,
) -> None:
    """
    Exportuje výsledky do XLSX.

    Sheets:
    - Candidates_All
    - Products
    - Components
    - Comparison
    - Excluded
    - Evidence
    - BOM_Options
    - Source_Checks
    - Final_Scoring_Weights
    - Legacy_Equivalence_Weights
    - Config
    """
    wb = openpyxl.load_workbook(template_path)

    registry_df = pd.DataFrame() if registry_df is None else registry_df.copy()
    products_df = pd.DataFrame() if products_df is None else products_df.copy()
    comparison_df = pd.DataFrame() if comparison_df is None else comparison_df.copy()
    excluded_df = pd.DataFrame() if excluded_df is None else excluded_df.copy()
    evidence_df = pd.DataFrame() if evidence_df is None else evidence_df.copy()
    bom_options_df = pd.DataFrame() if bom_options_df is None else bom_options_df.copy()
    components_df = pd.DataFrame() if components_df is None else components_df.copy()

    # Odstranit staré template listy, které už nemají být součástí hlavního scoringu.
    # Legacy equivalence data se exportuje explicitně do Legacy_Equivalence_Weights
    # s enabled=False.
    for obsolete_sheet in ["Equivalence_Weights"]:
        if obsolete_sheet in wb.sheetnames:
            ws_old = wb[obsolete_sheet]
            wb.remove(ws_old)

    # --- AUTO split base_set/component -> Components ---
    if not products_df.empty and components_df.empty:
        df = products_df.copy()

        if "candidate_type" in df.columns:
            cand = df["candidate_type"].astype(str).str.lower()
        else:
            cand = pd.Series([""] * len(df), index=df.index)

        if "complete_system" in df.columns:
            comp = df["complete_system"].astype(str).str.lower()
        else:
            comp = pd.Series([""] * len(df), index=df.index)

        is_component = (
            cand.isin(["base_set", "component"])
            | comp.str.contains("component/base-set", na=False)
            | comp.str.contains("component", na=False)
        )

        components_df = df[is_component].copy()
        products_df = df[~is_component].copy()

    def write_df(sheet_name: str, df: pd.DataFrame) -> None:
        # Přepiš sheet, aby v template nezůstávaly staré/hybridní hodnoty.
        if sheet_name in wb.sheetnames:
            ws_old = wb[sheet_name]
            wb.remove(ws_old)

        ws = wb.create_sheet(sheet_name)

        df = pd.DataFrame() if df is None else df.copy()

        # Hlavička
        ws.append([_sanitize_excel_string(str(c)) for c in df.columns])

        # Řádky
        for _, row in df.iterrows():
            ws.append([_to_excel_cell(v) for v in row.tolist()])

    write_df("Candidates_All", registry_df)
    write_df("Products", products_df)
    write_df("Components", components_df)
    write_df("Comparison", comparison_df)
    write_df("Excluded", excluded_df)
    write_df("Evidence", evidence_df)
    write_df("BOM_Options", bom_options_df)
    write_df("Source_Checks", _extract_source_checks(evidence_df))
    write_df("Final_Scoring_Weights", _extract_final_scoring_weights(cfg))
    write_df("Legacy_Equivalence_Weights", _extract_legacy_equivalence_weights(cfg))
    write_df("Config", _extract_config_sheet())
    write_df("Scoring_Field_Coverage", _scoring_field_coverage(products_df, comparison_df))

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)