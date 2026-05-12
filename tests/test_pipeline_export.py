import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pandas as pd

from src.config import default_config
from src.excel_export import export_excel
from src import pipeline


class _FakeConnectorYes:
    @staticmethod
    def extract_parameters(url):
        return {"flow_rate_lps": 0.6, "evidence": [("Flow", "0.6 l/s", url)]}

    @staticmethod
    def get_bom_options(url, params=None):
        return []


class _FakeDiscoverA:
    @staticmethod
    def discover_candidates(target_length_mm=1200, tolerance_mm=100):
        return ([{"manufacturer": "dallmer", "product_id": "d-1", "product_name": "A", "product_url": "https://a.example/p"}], [])


class _FakeDiscoverB:
    @staticmethod
    def discover_candidates(target_length_mm=1200, tolerance_mm=100):
        return ([{"manufacturer": "hansgrohe", "product_id": "h-1", "product_name": "B", "product_url": "https://b.example/p"}], [])


class _FakeViegaConnector:
    @staticmethod
    def extract_parameters(url):
        return {"flow_rate_lps": 0.7, "outlet_dn": "DN50", "material_detail": "Edelstahl 1.4301", "evidence": []}

    @staticmethod
    def get_bom_options(url, params=None):
        return []


class _FakeAcoConnector:
    @staticmethod
    def extract_parameters(url):
        if "57-128" in url or "ws25" in url:
            return {"flow_rate_lps": 0.62, "flow_rate_10mm_lps": 0.56, "flow_rate_20mm_lps": 0.62, "evidence": [("flow", "0,56/0,62 l/s", "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-c/")]}
        if "80-128" in url or "ws50" in url:
            return {"flow_rate_lps": 0.91, "flow_rate_10mm_lps": 0.72, "flow_rate_20mm_lps": 0.91, "evidence": [("flow", "0,72/0,91 l/s", "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-c/")]}
        if "public" in url:
            return {"flow_rate_lps": None, "din_en_1253_cert": None, "evidence": [("public", "official product page", "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/")]}
        return {"flow_rate_lps": 0.8, "evidence": [("splus", "official", "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-splus/")]}

    @staticmethod
    def get_bom_options(url, params=None):
        return []


class PipelineExportTests(unittest.TestCase):
    def _make_template(self, path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates_All"
        ws.append(["old"])
        ws.append(["stale"])
        for name in ["Excluded", "Products", "BOM_Options", "Components", "Evidence", "Comparison"]:
            x = wb.create_sheet(name)
            x.append(["old"])
            x.append(["stale"])
        wb.save(path)

    def _sheet_rows(self, path: Path, sheet: str):
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        return list(ws.iter_rows(values_only=True))

    def test_export_overwrites_candidates_all_with_latest_registry_only(self):
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "template.xlsx"
            out = Path(td) / "out.xlsx"
            self._make_template(template)

            first = pd.DataFrame([{"manufacturer": "Dallmer", "product_id": "d1"}])
            second = pd.DataFrame([{"manufacturer": "hansgrohe", "product_id": "h1"}])

            export_excel(template, out, default_config(), registry_df=first)
            export_excel(template, out, default_config(), registry_df=second)

            rows = self._sheet_rows(out, "Candidates_All")
            self.assertEqual(rows[0], ("manufacturer", "product_id"))
            self.assertEqual(rows[1], ("hansgrohe", "h1"))
            self.assertEqual(len(rows), 2)

    def test_export_overwrites_products_excluded_and_evidence_with_latest_run(self):
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "template.xlsx"
            out = Path(td) / "out.xlsx"
            self._make_template(template)

            export_excel(
                template,
                out,
                default_config(),
                registry_df=pd.DataFrame([{"manufacturer": "dallmer", "product_id": "old"}]),
                products_df=pd.DataFrame([{"manufacturer": "dallmer", "product_id": "old"}]),
                comparison_df=pd.DataFrame([{"manufacturer": "dallmer", "product_id": "old"}]),
                excluded_df=pd.DataFrame([{"manufacturer": "dallmer", "product_id": "old", "excluded_reason": "old"}]),
                evidence_df=pd.DataFrame([{"manufacturer": "dallmer", "product_id": "old", "label": "old", "source": "old"}]),
            )
            export_excel(
                template,
                out,
                default_config(),
                registry_df=pd.DataFrame([{"manufacturer": "hansgrohe", "product_id": "new"}]),
                products_df=pd.DataFrame([{"manufacturer": "hansgrohe", "product_id": "new"}]),
                comparison_df=pd.DataFrame([{"manufacturer": "hansgrohe", "product_id": "new"}]),
                excluded_df=pd.DataFrame([{"manufacturer": "aco", "product_id": "x1", "excluded_reason": "missing_flow"}]),
                evidence_df=pd.DataFrame([{"manufacturer": "hansgrohe", "product_id": "new", "label": "Flow", "source": "u"}]),
            )

            self.assertEqual(self._sheet_rows(out, "Products")[1], ("hansgrohe", "new"))
            self.assertEqual(self._sheet_rows(out, "Comparison")[1], ("hansgrohe", "new"))
            self.assertEqual(self._sheet_rows(out, "Excluded")[1], ("aco", "x1", "missing_flow"))
            self.assertEqual(self._sheet_rows(out, "Evidence")[1], ("hansgrohe", "new", "Flow", "u"))
            self.assertEqual(len(self._sheet_rows(out, "Products")), 2)
            self.assertEqual(len(self._sheet_rows(out, "Excluded")), 2)
            self.assertEqual(len(self._sheet_rows(out, "Evidence")), 2)

    def test_run_update_excludes_complete_system_no_and_normalizes_manufacturer(self):
        registry = pd.DataFrame(
            [
                {
                    "manufacturer": "Dallmer",
                    "product_id": "A1",
                    "product_name": "Accessory",
                    "product_url": "https://example.com/a1",
                    "candidate_type": "component",
                    "complete_system": "NO",
                    "excluded_reason": "complete_system_no",
                },
                {
                    "manufacturer": "dAllMer",
                    "product_id": "A2",
                    "product_name": "Drain",
                    "product_url": "https://example.com/a2",
                    "candidate_type": "drain",
                    "complete_system": "yes",
                },
            ]
        )
        with patch.dict(pipeline.CONNECTORS, {"dallmer": _FakeConnectorYes()}, clear=False):
            products, comparison, excluded, evidence, bom = pipeline.run_update(registry, default_config())

        self.assertEqual(products["manufacturer"].tolist(), ["dallmer"])
        self.assertEqual(products["product_id"].tolist(), ["A2"])
        self.assertEqual(comparison["product_id"].tolist(), ["A2"])
        self.assertEqual(excluded["product_id"].tolist(), ["A1"])
        self.assertEqual(excluded["excluded_reason"].tolist(), ["complete_system_no"])
        self.assertTrue((evidence["product_id"] == "A2").all())
        self.assertTrue(bom.empty)

    def test_run_discovery_respects_selected_connectors(self):
        with patch.dict(pipeline.CONNECTORS, {"dallmer": _FakeDiscoverA(), "hansgrohe": _FakeDiscoverB()}, clear=True):
            reg, dbg = pipeline.run_discovery(selected_connectors=["hansgrohe"])
        self.assertEqual(reg["manufacturer"].tolist(), ["hansgrohe"])
        self.assertTrue(dbg.empty)

    def test_run_update_respects_selected_connectors(self):
        registry = pd.DataFrame(
            [
                {"manufacturer": "dallmer", "product_id": "D1", "product_name": "D", "product_url": "https://d.example/p", "candidate_type": "drain", "complete_system": "yes"},
                {"manufacturer": "hansgrohe", "product_id": "H1", "product_name": "H", "product_url": "https://h.example/p", "candidate_type": "drain", "complete_system": "yes"},
            ]
        )
        with patch.dict(pipeline.CONNECTORS, {"dallmer": _FakeConnectorYes(), "hansgrohe": _FakeConnectorYes()}, clear=True):
            products, comparison, excluded, evidence, bom = pipeline.run_update(registry, default_config(), selected_connectors=["hansgrohe"])
        self.assertEqual(products["manufacturer"].tolist(), ["hansgrohe"])
        self.assertEqual(comparison["manufacturer"].tolist(), ["hansgrohe"])
        self.assertTrue(excluded.empty)
        self.assertTrue((evidence["manufacturer"] == "hansgrohe").all())
        self.assertTrue(bom.empty)

    def test_viega_lone_entities_remain_components_not_products(self):
        registry = pd.DataFrame(
            [
                {"manufacturer": "viega", "product_id": "v-498210", "product_name": "Advantix-Duschrinnen-Grundkörper 4982.10", "product_url": "https://v.example/4982-10.html", "candidate_type": "component", "complete_system": "yes", "system_role": "base_set", "discovery_seed_family": "advantix_line"},
                {"manufacturer": "viega", "product_id": "v-498294", "product_name": "Advantix-Duschrinnen-Geruchverschluss 4982.94", "product_url": "https://v.example/4982-94.html", "candidate_type": "component", "complete_system": "yes", "system_role": "base_set", "discovery_seed_family": "advantix_line"},
                {"manufacturer": "viega", "product_id": "v-493361", "product_name": "Advantix-Rost 4933.61", "product_url": "https://v.example/4933-61.html", "candidate_type": "component", "complete_system": "yes", "system_role": "cover", "discovery_seed_family": "advantix_line"},
                {"manufacturer": "viega", "product_id": "v-498291", "product_name": "Advantix-Verstellfußset 4982.91", "product_url": "https://v.example/4982-91.html", "candidate_type": "component", "complete_system": "yes", "system_role": "accessory", "discovery_seed_family": "advantix_line"},
            ]
        )
        with patch.dict(pipeline.CONNECTORS, {"viega": _FakeViegaConnector()}, clear=True):
            products, comparison, excluded, evidence, bom = pipeline.run_update(registry, default_config())
        self.assertFalse(products.empty)
        self.assertTrue((products["candidate_type"] == "component").all())
        self.assertTrue((products["promote_to_product"] == "no").all())
        self.assertFalse(comparison.empty)
        self.assertTrue(excluded.empty)
        self.assertIn("Viega promotion", evidence["label"].tolist())
        self.assertTrue(bom.empty)

    def test_viega_complete_assembly_promotes_body_to_product(self):
        registry = pd.DataFrame(
            [
                {"manufacturer": "viega", "product_id": "v-498210", "product_name": "Advantix-Duschrinnen-Grundkörper 4982.10", "product_url": "https://v.example/4982-10.html", "candidate_type": "component", "complete_system": "yes", "system_role": "base_set", "discovery_seed_family": "advantix_line"},
                {"manufacturer": "viega", "product_id": "v-498211", "product_name": "Advantix-Duschrinne 4982.11", "product_url": "https://v.example/4982-11.html", "candidate_type": "drain", "complete_system": "yes", "system_role": "complete_drain", "discovery_seed_family": "advantix_line"},
                {"manufacturer": "viega", "product_id": "v-493361", "product_name": "Advantix-Rost 4933.61", "product_url": "https://v.example/4982-61.html", "candidate_type": "component", "complete_system": "yes", "system_role": "cover", "discovery_seed_family": "advantix_line"},
            ]
        )
        with patch.dict(pipeline.CONNECTORS, {"viega": _FakeViegaConnector()}, clear=True):
            products, comparison, excluded, evidence, bom = pipeline.run_update(registry, default_config())
        self.assertGreaterEqual(len(products), 1)
        self.assertTrue((products["manufacturer"] == "viega").all())
        self.assertIn("yes", set(products["promote_to_product"].tolist()))
        promoted = products[products["promote_to_product"] == "yes"]
        self.assertTrue(all("rost" not in str(x).lower() for x in promoted["product_name"].tolist()))
        self.assertTrue(excluded.empty)
        self.assertTrue(bom.empty)

    def test_viega_badablauf_pages_are_drain_body_not_accessory(self):
        for name in [
            "Advantix Top-Badablauf 4914-20",
            "Advantix Top Badablauf 4914-20",
            "Advantix-Badablauf 4980-60",
            "Advantix-Badablauf 4980-61",
            "Advantix-Badablauf 4980-63",
            "Advantix Top-Bodenablauf 4914-11",
            "Advantix Top-Bodenablauf 4914-21",
            "Advantix-Bodenablauf-Grundkörper 4951-15",
            "Advantix-Bodenablauf-Grundkörper 4955-15",
            "Advantix-Bodenablauf-Grundkörper 4955-25",
        ]:
            role = pipeline._infer_viega_role({"system_role": "accessory", "product_name": name, "product_url": f"https://v.example/{name.replace(' ', '-')}.html"})
            self.assertEqual(role, "base_set")

    def test_viega_drain_body_pages_use_incomplete_assembly_not_non_promotable_accessory(self):
        registry = pd.DataFrame(
            [
                {"manufacturer": "viega", "product_id": "v-491420", "product_name": "Advantix Top-Badablauf 4914-20", "product_url": "https://v.example/4914-20.html", "candidate_type": "component", "complete_system": "yes", "system_role": "accessory", "discovery_seed_family": "advantix_floor"},
                {"manufacturer": "viega", "product_id": "v-498060", "product_name": "Advantix-Badablauf 4980-60", "product_url": "https://v.example/4980-60.html", "candidate_type": "component", "complete_system": "yes", "system_role": "accessory", "discovery_seed_family": "advantix_floor"},
            ]
        )
        with patch.dict(pipeline.CONNECTORS, {"viega": _FakeViegaConnector()}, clear=True):
            products, comparison, excluded, evidence, bom = pipeline.run_update(registry, default_config())
        self.assertFalse(products.empty)
        self.assertTrue((products["promote_to_product"] == "no").all())
        self.assertTrue((products["why_not_product_reason"] == "incomplete_assembly").all())
        self.assertNotIn("non_promotable_accessory", set(products["why_not_product_reason"].tolist()))
        self.assertTrue(excluded.empty)
        self.assertTrue(bom.empty)

    def test_viega_explicit_override_ids_force_base_set_incomplete_assembly(self):
        rows = []
        for pid, name in [
            ("viega-491420", "Advantix Top-Badablauf 4914-20"),
            ("viega-498060", "Advantix-Badablauf 4980-60"),
            ("viega-498061", "Advantix-Badablauf 4980-61"),
            ("viega-498063", "Advantix-Badablauf 4980-63"),
            ("viega-495120", "Advantix-Bodenablauf 4951-20"),
            ("viega-495115", "Advantix-Bodenablauf-Grundkörper 4951-15"),
            ("viega-495515", "Advantix-Bodenablauf-Grundkörper 4955-15"),
            ("viega-495525", "Advantix-Bodenablauf-Grundkörper 4955-25"),
            ("viega-491411", "Advantix Top-Bodenablauf 4914-11"),
            ("viega-491421", "Advantix Top-Bodenablauf 4914-21"),
        ]:
            rows.append(
                {
                    "manufacturer": "viega",
                    "product_id": pid,
                    "product_name": name,
                    "product_url": f"https://v.example/{pid}.html",
                    "candidate_type": "component",
                    "complete_system": "yes",
                    "system_role": "accessory",
                    "discovery_seed_family": "advantix_floor",
                }
            )
        registry = pd.DataFrame(rows)
        with patch.dict(pipeline.CONNECTORS, {"viega": _FakeViegaConnector()}, clear=True):
            products, comparison, excluded, evidence, bom = pipeline.run_update(registry, default_config())
        self.assertEqual(len(products), 10)
        self.assertTrue((products["promote_to_product"] == "no").all())
        self.assertTrue((products["why_not_product_reason"] == "incomplete_assembly").all())
        self.assertNotIn("non_promotable_accessory", set(products["why_not_product_reason"].tolist()))
        self.assertTrue(excluded.empty)
        self.assertTrue(bom.empty)

    def test_aco_showerdrain_c_low_and_standard_variants_not_mixed(self):
        registry = pd.DataFrame(
            [
                {"manufacturer": "aco", "product_id": "a-low", "product_name": "ShowerDrain C WS25", "product_url": "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-c/rinnenkoerper-57-128-mm-ws25/", "candidate_type": "drain", "complete_system": "yes"},
                {"manufacturer": "aco", "product_id": "a-std", "product_name": "ShowerDrain C WS50", "product_url": "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-c/rinnenkoerper-80-128-mm-ws50/", "candidate_type": "drain", "complete_system": "yes"},
            ]
        )
        with patch.dict(pipeline.CONNECTORS, {"aco": _FakeAcoConnector()}, clear=True):
            products, comparison, _, _, _ = pipeline.run_update(registry, default_config())
        low = products[products["product_id"] == "a-low"].iloc[0]
        std = products[products["product_id"] == "a-std"].iloc[0]
        self.assertEqual(low["flow_rate_20mm_lps"], 0.62)
        self.assertNotEqual(low["flow_rate_lps"], 0.91)
        self.assertEqual(std["flow_rate_20mm_lps"], 0.91)
        self.assertEqual(set(comparison["product_id"].tolist()), {"a-low", "a-std"})

    def test_aco_comparison_excludes_component_only_rows(self):
        registry = pd.DataFrame(
            [
                {"manufacturer": "aco", "product_id": "a-splus", "product_name": "ShowerDrain S+", "product_url": "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-splus/", "candidate_type": "drain", "complete_system": "yes"},
                {"manufacturer": "aco", "product_id": "a-mplus-body", "product_name": "ShowerDrain M+ body", "product_url": "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-mplus/", "candidate_type": "component", "system_role": "base_set", "complete_system": "yes"},
                {"manufacturer": "aco", "product_id": "a-mplus-grate", "product_name": "ShowerDrain M+ grate", "product_url": "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/aco-showerdrain-mplus/", "candidate_type": "component", "system_role": "grate", "complete_system": "yes"},
            ]
        )
        with patch.dict(pipeline.CONNECTORS, {"aco": _FakeAcoConnector()}, clear=True):
            products, comparison, excluded, evidence, _ = pipeline.run_update(registry, default_config())
        self.assertEqual(set(products["product_id"].tolist()), {"a-splus", "a-mplus-body", "a-mplus-grate"})
        self.assertEqual(comparison["product_id"].tolist(), ["a-splus"])
        self.assertTrue(excluded.empty)
        self.assertTrue(all(("aco-haustechnik.de" in s) for s in evidence["source"].tolist()))

    def test_aco_public_products_not_excluded_when_flow_missing(self):
        registry = pd.DataFrame(
            [{"manufacturer": "aco", "product_id": "a-public80", "product_name": "Public 80", "product_url": "https://www.aco-haustechnik.de/produkte/badentwaesserung/duschrinnen/public-80/", "candidate_type": "drain", "complete_system": "yes"}]
        )
        with patch.dict(pipeline.CONNECTORS, {"aco": _FakeAcoConnector()}, clear=True):
            products, comparison, excluded, _, _ = pipeline.run_update(registry, default_config())
        self.assertEqual(products["product_id"].tolist(), ["a-public80"])
        self.assertEqual(comparison["product_id"].tolist(), ["a-public80"])
        self.assertTrue(excluded.empty)


if __name__ == "__main__":
    unittest.main()
